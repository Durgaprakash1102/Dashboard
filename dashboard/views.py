from django.http import JsonResponse
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from .models import CustomUser
from django.core.mail import send_mail
import random

def signup(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        role = request.POST['role']
        email = request.POST['email']
        otp = random.randint(100000, 999999)
        
        send_mail(
            'Your OTP',
            f'Your OTP is {otp}',
            'from@example.com',
            [email],
            fail_silently=False,
        )

        request.session['otp'] = otp
        request.session['username'] = username
        request.session['password'] = password
        request.session['role'] = role
        request.session['email'] = email
        return redirect('verify_otp')

    return render(request, 'signup.html')

def verify_otp(request):
    if request.method == 'POST':
        entered_otp = request.POST['otp']
        if entered_otp == str(request.session.get('otp')):
            user = CustomUser.objects.create_user(
                username=request.session['username'],
                password=request.session['password'],
                role=request.session['role'],
                email=request.session['email']
            )
            login(request, user)
            return redirect('dashboard')
    
    return render(request, 'verify_otp.html')

def login_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            return redirect('dashboard')
        else:
            messages.error(request, 'Invalid username or password.')
    
    return render(request, 'login.html')


# def dashboard(request):
#     if request.user.is_authenticated:
#         if request.user.role == 'team_1':
#             return redirect('team_1_dashboard')
#         elif request.user.role == 'team_2':
#             return redirect('team_2_dashboard')
#         elif request.user.role == 'team_3':
#             return redirect('team_3_dashboard')
#         elif request.user.role == 'team_4':
#             return redirect('critical_loans')
#         elif request.user.role == 'team_5':
#             return redirect('team_4_dashboard')
#         elif request.user.role == 'super_admin':
#             return redirect('super_admin_dashboard')
#         elif request.user.role=='':
#             return redirect('login')
#     return redirect('login')

from django.shortcuts import redirect

def dashboard(request):
    # Check if the user is authenticated
    if request.user.is_authenticated:
        # Redirect based on user role
        if request.user.is_superuser:  # Check if the user is a superuser
            return redirect('super_admin_dashboard')
        elif request.user.role == 'team_1':
            return redirect('team_1_dashboard')
        elif request.user.role == 'team_2':
            return redirect('team_2_dashboard')
        elif request.user.role == 'team_3':
            return redirect('team_3_dashboard')
        elif request.user.role == 'team_4':
            return redirect('critical_loans')
        elif request.user.role == 'super_admin':
            return redirect('team_4_dashboard')
        elif request.user.role == '':
            return redirect('login')
    
    # If the user is not authenticated or role doesn't match
    return redirect('login')


from django.shortcuts import render, redirect, get_object_or_404
from .models import Customer

def team_1_dashboard(request):
    customers = Customer.objects.all()
    
    # Separate lists for interested and not interested customers
    interested_customers = customers.filter(is_interested=True)
    not_interested_customers = customers.filter(is_not_interested=True)

    return render(request, 'team_1_dashboard.html', {
        'customers': customers,
        'interested_customers': interested_customers,
        'not_interested_customers': not_interested_customers,
    })


def update_interest_status(request, customer_id):
    customer = get_object_or_404(Customer, id=customer_id)
    
    if request.method == 'POST':
        action = request.POST.get('interest_action', '')

        if action == 'interested':
            customer.is_interested = True
            customer.is_not_interested = False
            customer.reason = "Move Forward"
        
        elif action == 'not_interested':
            customer.is_interested = False
            customer.is_not_interested = True
            customer.reason = request.POST.get('reason', '')

        customer.save()
        return redirect('team_1_dashboard')

    return render(request, 'team_1_dashboard.html', {'customers': Customer.objects.all()})

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from .models import Customer

def team_2_dashboard(request):
    # Fetch customers with different eligibility statuses
    interested_customers = Customer.objects.filter(is_interested=True, is_eligible=None, is_not_eligible=None)
    eligible_customers = Customer.objects.filter(is_eligible=True, is_interested=True)
    not_eligible_customers = Customer.objects.filter(is_not_eligible=True, is_interested=True)

    return render(request, 'team_2_dashboard.html', {
        'interested_customers': interested_customers,
        'eligible_customers': eligible_customers,
        'not_eligible_customers': not_eligible_customers,
    })

def update_eligibility_status(request, customer_id):
    if request.method == 'POST':
        action = request.POST.get('eligibility_action')
        customer = get_object_or_404(Customer, id=customer_id)

        # Set the appropriate status based on the action
        if action == 'eligible':
            customer.is_eligible = True
            customer.is_not_eligible = False
        elif action == 'not_eligible':
            customer.is_eligible = False
            customer.is_not_eligible = True

        customer.save()

        # Add a success message
        messages.success(request, f'Customer {customer.name} has been marked as {action.replace("_", " ")}.')

        return redirect('team_2_dashboard')

    return redirect('team_2_dashboard')

from django.shortcuts import render, get_object_or_404, redirect
from .models import Loan
from datetime import date  

def team_3_dashboard(request):
    query = request.GET.get('q')

    # Loans where EMI is not cleared or blank
    pending_loans = Loan.objects.filter(emi_cleared__in=[None, False])

    # Loans where EMI is cleared
    cleared_loans = Loan.objects.filter(emi_cleared=True)

    if query:
        pending_loans = pending_loans.filter(loan_id__icontains=query)
        cleared_loans = cleared_loans.filter(loan_id__icontains=query)

    # Ensure to check and update critical status for all loans
    for loan in pending_loans.union(cleared_loans):
        loan.check_and_update_critical_status()

    return render(request, 'team_3_dashboard.html', {
        'pending_loans': pending_loans,
        'cleared_loans': cleared_loans
    })

def update_emi_status(request, loan_id):
    loan = get_object_or_404(Loan, loan_id=loan_id)

    if request.method == 'POST':
        emi_cleared = request.POST.get('emi_cleared')
        due_date_pending_reason = request.POST.get('due_date_pending_reason', '')

        if emi_cleared == 'true':
            loan.emi_cleared = True
            loan.due_date_pending_reason = ''
        else:
            loan.emi_cleared = False
            loan.due_date_pending_reason = due_date_pending_reason

        loan.save()
        loan.check_and_update_critical_status()

        return redirect('team_3_dashboard')

def extend_due_date(request, loan_id):
    loan = get_object_or_404(Loan, loan_id=loan_id)

    if request.method == 'POST':
        extended_due_date = request.POST.get('extended_due_date')
        if extended_due_date:
            loan.extended_due_date = date.fromisoformat(extended_due_date)
            loan.save()
            loan.check_and_update_critical_status()

        return redirect('team_3_dashboard')
    
from django.shortcuts import render, redirect, get_object_or_404
from .models import Loan

def critical_loans(request):
    # Fetch only loans with critical status and pending EMI
    critical_loans = Loan.objects.filter(
        critical_loan_status='CRITICAL',
        emi_cleared=False  # Exclude loans where EMI is cleared
    )
    return render(request, 'critical_loans.html', {'critical_loans': critical_loans})

def update_visit_details(request, loan_id):
    loan = get_object_or_404(Loan, id=loan_id)
    
    if request.method == 'POST':
        visit_remarks = request.POST.get('visit_remarks', '')
        visit_date = request.POST.get('visit_date', loan.visit_date)
        revisit_date = request.POST.get('revisit_date', None)
        revisit_reason = request.POST.get('revisit_reason', None)

        # Update visit details in the loan record
        loan.visit_remarks = visit_remarks
        loan.visit_date = visit_date
        
        if revisit_date and revisit_reason:
            loan.revisit_date = revisit_date
            loan.revisit_reason = revisit_reason

        loan.save()
        return redirect('critical_loans')

def mark_emi_cleared(request, loan_id):
    loan = get_object_or_404(Loan, id=loan_id)
    
    # Mark the EMI as cleared
    loan.emi_cleared = True
    
    # Change critical loan status to NORMAL
    loan.critical_loan_status = 'NORMAL'
    
    loan.save()
    
    # Redirect to critical loans page where cleared loans won't be shown
    return redirect('critical_loans')



def team_4_dashboard(request):
    return render(request, 'team_4_dashboard.html')

import openpyxl
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import Loan
from django.utils.dateparse import parse_date

def upload_loan_details(request):
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        
        # Check if the file is an Excel file
        if file.name.endswith('.xlsx'):
            try:
                wb = openpyxl.load_workbook(file)
                sheet = wb.active

                # Validate the header structure in Excel
                expected_headers = ['name', 'loan_id', 'phone', 'present_loan_amount', 
                                    'tenure', 'total_monthly_emi_amount', 
                                    'emi_due_date', 'emi_cleared', 
                                    'extended_due_date', 'critical_loan_status', 
                                    'visit_date', 'visit_remarks', 
                                    'revisit_date', 'revisit_reason', 
                                    'due_date_pending_reason']
                
                actual_headers = [cell.value for cell in sheet[1]]

                # Check if the headers match the expected structure
                if actual_headers[:7] != expected_headers[:7]:
                    messages.error(request, "Invalid Excel structure. Please follow the correct format.")
                    return redirect('upload_loan_details')

                # Iterate through the rows, skipping the header
                for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    name = row[0]
                    loan_id = row[1]
                    phone = row[2]
                    present_loan_amount = row[3]
                    tenure = row[4]
                    total_monthly_emi_amount = row[5]
                    emi_due_date = row[6]
                    
                    # Optional fields
                    emi_cleared = row[7] if len(row) > 7 else None
                    extended_due_date = row[8] if len(row) > 8 else None
                    critical_loan_status = row[9] if len(row) > 9 else None
                    visit_date = row[10] if len(row) > 10 else None
                    visit_remarks = row[11] if len(row) > 11 else None
                    revisit_date = row[12] if len(row) > 12 else None
                    revisit_reason = row[13] if len(row) > 13 else None
                    due_date_pending_reason = row[14] if len(row) > 14 else None

                    # Parse dates if present
                    emi_due_date = parse_date(emi_due_date) if emi_due_date else None
                    extended_due_date = parse_date(extended_due_date) if extended_due_date else None
                    visit_date = parse_date(visit_date) if visit_date else None
                    revisit_date = parse_date(revisit_date) if revisit_date else None

                    # Ensure required fields are present, use default values for missing non-null fields
                    if all([name, loan_id, phone, present_loan_amount, tenure, total_monthly_emi_amount, emi_due_date]):
                        # Default value handling
                        emi_cleared = bool(emi_cleared) if emi_cleared is not None else False
                        critical_loan_status = critical_loan_status if critical_loan_status else Loan.CriticalStatus.NORMAL

                        # Create and save loan data
                        Loan.objects.create(
                            name=name,
                            loan_id=loan_id,
                            phone=phone,
                            present_loan_amount=present_loan_amount,
                            tenure=tenure,
                            total_monthly_emi_amount=total_monthly_emi_amount,
                            emi_due_date=emi_due_date,
                            emi_cleared=emi_cleared,
                            extended_due_date=extended_due_date,
                            critical_loan_status=critical_loan_status,
                            visit_date=visit_date,
                            visit_remarks=visit_remarks,
                            revisit_date=revisit_date,
                            revisit_reason=revisit_reason,
                            due_date_pending_reason=due_date_pending_reason
                        )
                    else:
                        # If any required field is missing
                        messages.error(request, f"Missing required fields in row {row_idx}.")
                        return redirect('upload_loan_details')

                messages.success(request, "Loan details successfully uploaded.")
                return redirect('team_4_dashboard')
            except Exception as e:
                messages.error(request, f"Error processing file: {e}")
                return redirect('upload_loan_details')
        else:
            messages.error(request, "Please upload a valid Excel (.xlsx) file.")
            return redirect('upload_loan_details')

    return render(request, 'upload_loan_details.html')

import openpyxl
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import Customer
from datetime import datetime

def parse_date(date_str):
    """Attempt to parse a date string in multiple formats."""
    formats = ['%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%d.%m.%Y']  # Add more formats as needed
    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
    return None  # Return None if all formats fail

def upload_customer_details(request):
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        
        # Check if the file is an Excel file
        if file.name.endswith('.xlsx'):
            try:
                wb = openpyxl.load_workbook(file)
                sheet = wb.active

                # Validate the header structure in Excel
                expected_headers = ['name', 'phone', 'email', 'dob', 'gender', 'address', 
                                    'is_interested', 'is_not_interested', 'is_eligible', 
                                    'is_not_eligible', 'reason']
                actual_headers = [cell.value for cell in sheet[1]]

                # Check if the headers match the expected structure
                if actual_headers[:6] != expected_headers[:6]:
                    messages.error(request, "Invalid Excel structure. Please follow the correct format.")
                    return redirect('upload_customer_details')

                # Iterate through the rows, skipping the header
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    name = row[0]
                    phone = row[1]
                    email = row[2]
                    dob = parse_date(row[3])  # Convert date using the updated function
                    gender = row[4]
                    address = row[5]

                    # Optional fields
                    is_interested = row[6] if len(row) > 6 else None
                    is_not_interested = row[7] if len(row) > 7 else None
                    is_eligible = row[8] if len(row) > 8 else None
                    is_not_eligible = row[9] if len(row) > 9 else None
                    reason = row[10] if len(row) > 10 else None

                    # Ensure required fields are present
                    if all([name, phone, email, dob, gender, address]):
                        # Create and save customer data
                        Customer.objects.create(
                            name=name,
                            phone=phone,
                            email=email,
                            dob=dob,
                            gender=gender,
                            address=address,
                            is_interested=is_interested,
                            is_not_interested=is_not_interested,
                            is_eligible=is_eligible,
                            is_not_eligible=is_not_eligible,
                            reason=reason
                        )
                    else:
                        # If any required field is missing
                        messages.error(request, f"Missing required fields in row {sheet.iter_rows(min_row=2).index(row) + 2}.")
                        return redirect('upload_customer_details')

                messages.success(request, "Customer details successfully uploaded.")
                return redirect('team_4_dashboard')
            except Exception as e:
                messages.error(request, f"Error processing file: {e}")
                return redirect('upload_customer_details')
        else:
            messages.error(request, "Please upload a valid Excel (.xlsx) file.")
            return redirect('upload_customer_details')

    return render(request, 'upload_customer_details.html')

def manager_view_customers(request):
    customers = Customer.objects.all()
    return render(request, 'manager_view_customers.html',{'customers': customers})

def manager_view_loans(request):
    loans = Loan.objects.all()
    return render(request, 'manager_view_loans.html',{'loans': loans})

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from .models import CustomUser

def manager_view_users(request):
    # Filter users belonging to team_1, team_2, team_3, and team_4
    users = CustomUser.objects.filter(role__in=['team_1', 'team_2', 'team_3', 'team_4'])
    return render(request, 'manager_view_users.html', {'users': users})

# View to delete a user
def manager_delete_user(request, user_id):
    user = get_object_or_404(CustomUser, id=user_id)  # Get user or return 404
    if user.is_superuser:
        messages.error(request, 'Cannot delete super users.')
    else:
        user.delete()  # Delete the user if not super admin
        messages.success(request, 'User deleted successfully.')
    return redirect('manager_view_users')  # Redirect to the manager view users page

from django.contrib import messages
from django.shortcuts import render, redirect
from django.contrib.auth.hashers import make_password
from .models import CustomUser

def manager_create_user(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        role = request.POST['role']
        email = request.POST['email']
        
        # Validate the role (managers can only assign team_1 to team_4 roles)
        valid_roles = ['team_1', 'team_2', 'team_3', 'team_4']
        if role not in valid_roles:
            messages.error(request, 'Invalid role selected. Managers can only create users for team_1 to team_4.')
            return render(request, 'manager_create_user.html')  # Return the form if role is invalid

        # Create user with hashed password and store the original password
        user = CustomUser(
            username=username,
            email=email,
            role=role,
            password=make_password(password),  # Hash the password
            original_password=password  # Store the original password temporarily
        )
        user.save()  # Save the user to the database
        messages.success(request, 'User created successfully.')
        return redirect('team_4_dashboard')

    return render(request, 'manager_create_user.html')  # Render form to create a new user








from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from .models import CustomUser, Customer, Loan
from django.contrib.auth.hashers import make_password

def super_admin_dashboard(request):
    # Fetch all customers, loans, and users
    customers = Customer.objects.all()
    loans = Loan.objects.all()
    users = CustomUser.objects.all()  # Use CustomUser instead of User
    return render(request, 'superadmin_dashboard.html', {
        'customers': customers,
        'loans': loans,
        'users': users
    })

# View to delete a user
def delete_user(request, user_id):
    user = get_object_or_404(CustomUser, id=user_id)  # Get user or return 404
    if user.is_superuser:
        messages.error(request, 'Cannot delete super users.')
    else:
        user.delete()  # Delete the user if not super admin
        messages.success(request, 'User deleted successfully.')
    return redirect('super_admin_dashboard')

from django.contrib import messages
from django.contrib.auth.hashers import make_password
from django.shortcuts import render, redirect
from .models import CustomUser  # Ensure this imports your CustomUser model

# View to create a new user
from django.contrib import messages
from django.shortcuts import render, redirect
from django.contrib.auth.hashers import make_password
from .models import CustomUser

def create_user(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        role = request.POST['role']
        email = request.POST['email']
        
        # Validate the role
        valid_roles = ['team_1', 'team_2', 'team_3', 'team_4', 'super_admin']
        if role not in valid_roles:
            messages.error(request, 'Invalid role selected.')
            return render(request, 'create_user.html')  # Return the same form if role is invalid

        # Create user with hashed password and store the original password
        user = CustomUser(
            username=username,
            email=email,
            role=role,
            password=make_password(password),  # Hash the password
            original_password=password  # Store the original password temporarily
        )
        user.save()  # Save the user to the database
        messages.success(request, 'User created successfully.')
        return redirect('super_admin_dashboard')

    return render(request, 'create_user.html')  # Render form to create a new user


# View to display all customers
def view_customers(request):
    customers = Customer.objects.all()
    return render(request, 'view_customers.html', {'customers': customers})

# View to display all loans
def view_loans(request):
    loans = Loan.objects.all()
    return render(request, 'view_loans.html', {'loans': loans})

# View to display all users
def view_users(request):
    users = CustomUser.objects.all()  # Use CustomUser instead of User
    return render(request, 'view_users.html', {'users': users})

def logout_view(request):
    logout(request)
    return redirect('login')

# Home page view
def home(request):
    return render(request, 'home.html')



from django.shortcuts import render

def popup_view(request):
    return render(request, 'popup.html')
